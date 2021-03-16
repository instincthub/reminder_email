#! python3
# Send Email reminder based on their status on the spreadsheet

import openpyxl
import smtplib

# TODO: Check each member's payment status.

wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb['Sheet1']

# Target row and column
lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value

unpaidMember = {}

for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMember[name] = email


# TODO: Log in to email account.
domain = 'smtp.gmail.com'
port = 587
input_pw = input('Email Password:')

# Try to login to the server and send our email

try:
    server = smtplib.SMTP(domain, port)
    server.ehlo()
    server.starttls()
    server.login('skills@instincthub.com', input_pw)

    # TODO: Send out reminder emails.
    for name, email in unpaidMember.items():
        body = f'Subject: {latestMonth} dues  unpaid \nDear {name}, \nRecord shows that you have not paid dues for {latestMonth}. \nPlease make payment as soon as possible. \n\nThank you'

        print(f'Sending Email to {name}...')

        sendEmail = server.sendmail('skills@instincthub.com', email, body)

        if sendEmail != {}:
            print(f'There is a problem sending email to {email}: {sendEmail}')

except Exception as e:
    # Print error
    print(e)

finally:
    server.quit()